#!/usr/bin/env python3
"""
report_builder.py
-----------------
Pure logic module — no GUI, no Flask, no tkinter.
Responsible for:
  1. Managing the master server list (master_servers.json)
  2. Matching incoming JSON data against that list
  3. Building and styling the Excel workbook
  4. Returning the finished file as raw bytes (so Flask can send it as a download)
  5. Writing a JSON run-log for every report generated

This module is intentionally decoupled from the web layer so it can be tested
or reused independently of Flask.
"""

import json
import math
import re
import traceback
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.comments import Comment
from pathlib import Path
from datetime import datetime, timezone
import io

# ─────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────

import os
# Name stamped into Excel cell comments (e.g. on multi-drive disk breakdowns)
OWNER_TAG = os.getenv("REPORT_AUTHOR", "Victor Ohiovbeunu")

# /app/data is a Docker named volume — data here persists across container restarts.
# When running locally (not in Docker) this path is still created, but the volume
# mapping in docker-compose.yml is what makes it truly persistent.
DATA_DIR    = Path("/app/data")
DATA_DIR.mkdir(parents=True, exist_ok=True)

MASTER_FILE = DATA_DIR / "master_servers.json"   # canonical ordered server list
LOG_DIR     = DATA_DIR / "logs"                  # one JSON log file per report run
LOG_DIR.mkdir(parents=True, exist_ok=True)

# ─────────────────────────────────────────────
# DEFAULT MASTER SERVER LIST
# ─────────────────────────────────────────────
# This list is written to master_servers.json on the FIRST run only (i.e. when the
# volume is empty / the file doesn't exist yet).  After that, the file on disk is
# the source of truth — you can edit it directly to reorder or add servers.
# Format: "ServerName : IP"  (the space-colon-space pattern is enforced by norm_name)
DEFAULT_MASTER = [
    {"name": "ADSERVER : 10.4.0.7"},
    {"name": "AutoDeployVM : 10.4.0.4"},
    {"name": "AutoDeployVM-2 : 10.4.0.28"},
    {"name": "AutoDeployVMTest : 10.4.0.4"},
    {"name": "AV-VM : 10.5.0.18"},
    {"name": "BackupAD : 10.4.0.8"},
    {"name": "BankoneAppStag2 : 172.1.4.12"},
    {"name": "bankonere000007 : 10.1.2.46"},
    {"name": "BankOneReportVM : 10.1.2.5"},
    {"name": "bankone-s000001 : 10.1.4.22"},
    {"name": "bankone-s000002 : 10.1.4.23"},
    {"name": "bankone-s00037Y : 10.1.4.17"},
    {"name": "blendappv000001 : 10.2.2.9"},
    {"name": "blendappv000003 : 10.2.2.11"},
    {"name": "blendappv000004 : 10.2.2.13"},
    {"name": "BLENDLIVEAPP012 : 10.2.2.8"},
    {"name": "BLENDLIVEAPP03 : 10.2.2.10"},
    {"name": "BLENDLIVEWEB01 : 10.2.1.6"},
    {"name": "blendwebv000001 : 10.2.1.10"},
    {"name": "blendwebv000003 : 10.2.1.11"},
    {"name": "BUILD-SERVER: 10.4.0.15"},
    {"name": "clusterap00000D : 10.2.2.55"},
    {"name": "Cluster-App-01 : 10.1.2.25"},
    {"name": "Cluster-App-02 : 10.1.2.27"},
    {"name": "Cluster-App-Dev : 10.1.4.12"},
    {"name": "coreapp02000001 : 10.1.2.35"},
    {"name": "coreapp3-000002 : 10.1.2.37"},
    {"name": "CoreApp02-VM : 10.1.2.6"},
    {"name": "CoreApp03-VM : 10.1.2.16"},
    {"name": "coreapp03000006 : 10.1.2.34"},
    {"name": "Coreapp2-000000 : 10.1.2.31"},
    {"name": "CoreWeb : 10.1.1.25"},
    {"name": "coreweb01000009 : 10.1.1.27"},
    {"name": "coreweb0100000A : 10.1.1.32"},
    {"name": "ELASTICSEARCH-VM : 10.5.0.6"},
    {"name": "ELASTICSEARCH-VM2: 10.5.0.17"},
    {"name": "FGRIDINTAPPVM-2 : 10.1.2.22"},
    {"name": "fingridap00000F : 10.1.2.57"},
    {"name": "fingridap00000J : 10.1.2.53"},
    {"name": "fingridapMIRPD8 : 10.1.2.62"},
    {"name": "FinGridApp01-1 : 10.1.2.13"},
    {"name": "FinGridApp01-2 : 10.1.2.14"},
    {"name": "FinGridApp02-2 : 10.1.2.7"},
    {"name": "FinGridApp02-3 : 10.1.2.8"},
    {"name": "FinGridApp02-5 : 10.1.2.18"},
    {"name": "FinGridApp03-1 : 10.1.2.9"},
    {"name": "FinGridApp03-3 : 10.1.2.70"},
    {"name": "FinGridApp03-4 : 10.1.2.12"},
    {"name": "FinGridAPP03-5 : 10.1.2.15"},
    {"name": "fingridapQNKSR5 : 10.1.2.28"},
    {"name": "FINGRIDDB02 : 10.1.3.7"},
    {"name": "FinGridDB03 : 10.1.3.8"},
    {"name": "FINGRIDINTAPPVM : 10.1.2.11"},
    {"name": "FinGridStagApp1 : 10.1.4.12"},
    {"name": "FINGRIDSTAGDB03 : 10.1.4.9"},
    {"name": "FinGridSTAGIB : 10.1.4.4"},
    {"name": "FinGridStagWeb1 : 10.1.4.13"},
    {"name": "FinGridWEB02-1 : 10.1.1.5"},
    {"name": "FinGridWeb02-2 : 10.1.1.8"},
    {"name": "FINGRIDWEB03-02 : 10.1.1.4"},
    {"name": "FINGRIDWEB03-1 : 10.1.1.12"},
    {"name": "FinGridWEB3-02R: 10.2.1.4"},
    {"name": "GRAFANAVM : 10.5.0.10"},
    {"name": "GRAFANAVM : 10.5.0.7"},
    {"name": "GRAYLOG-VM : 10.5.0.12"},
    {"name": "GRAYLOG-VM2 : 10.5.0.13"},
    {"name": "gridapp 0000009 : 10.1.2.40"},
    {"name": "iBANK-VM : 10.1.1.16"},
    {"name": "ibank-web000000 : 10.1.1.6"},
    {"name": "ibank-web000005 : 10.1.1.9"},
    {"name": "mybankone000002 : 10.1.1.24"},
    {"name": "mybankone000004 : 10.1.1.22"},
    {"name": "mybankoneCR86QJ : 10.1.1.13"},
    {"name": "mybankone1TROCT : 10.1.1.11"},
    {"name": "mybankone00000D : 10.1.1.33"},
    {"name": "NUGGET-DEV-DB : 10.4.0.245"},
    {"name": "QoreApp : 10.1.2.61"},
    {"name": "qoreapp030000003 : 10.1.2.39"},
    {"name": "Qore-switch-Pro : 10.1.2.7"},
    {"name": "RecovaLiveWeb1 : 10.7.1.4"},
    {"name": "RecovaBankone-D : 10.1.3.13"},
    {"name": "TBPRIMEAPP01 : 10.2.2.5"},
    {"name": "TEAMCITY-VM : 10.5.0.23"},
]

# ─────────────────────────────────────────────
# EXCEL STYLE OBJECTS
# ─────────────────────────────────────────────
# Cell background colours applied based on utilisation thresholds:
#   90–95%  → Yellow  (watch)
#   96–98%  → Orange  (concern)
#   99–100% → Red     (critical)
FILL_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
FILL_ORANGE = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
FILL_RED    = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Border styles used on the tables:
#   THIN_SIDE  → thin black grid lines inside the table
#   THICK_BLUE → thick blue outer frame around each table block
THIN_SIDE  = Side(border_style="thin",  color="000000")
THICK_BLUE = Side(border_style="thick", color="0000FF")


# ─────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────

def _seed_master():
    """
    Write DEFAULT_MASTER to disk if the master file doesn't exist yet.
    This only runs once — on the very first report generation after a fresh
    Docker volume is created.  After that, the on-disk file is never overwritten
    by this function.
    """
    if not MASTER_FILE.exists():
        with open(MASTER_FILE, "w", encoding="utf-8") as f:
            json.dump(DEFAULT_MASTER, f, indent=2, ensure_ascii=False)


def ceil_or_none(v):
    """
    Safely convert a value to an integer, rounding UP (ceiling).
    Returns None if the value is missing, empty, or non-numeric.

    We use ceiling (not rounding) so resource utilisation is never
    under-reported — e.g. 90.1% becomes 91, never 90.
    """
    if v is None or v == "":
        return None
    try:
        return math.ceil(float(v))
    except Exception:
        return None  # non-numeric string → silently ignore


def norm_name(n):
    """
    Normalise a server name string for fuzzy matching.

    The AI chatbot may produce names with slightly different spacing or
    capitalisation compared to what's stored in the master list.
    This function collapses all whitespace, standardises the ' : ' separator,
    and lowercases everything so that:
        "ADSERVER:10.4.0.7"  ==  "ADSERVER : 10.4.0.7"  ==  "adserver : 10.4.0.7"

    The normalised key is ONLY used for matching — the original display name
    from the master list is always used in the actual report.
    """
    if not n:
        return ""
    s = " ".join(str(n).split())     # collapse any internal whitespace runs
    if ":" in s:
        s = re.sub(r"\s*:\s*", " : ", s)  # standardise spacing around colon
    return s.strip().lower()


def pick_fill(v):
    """
    Return the correct PatternFill for a given utilisation percentage,
    or None if the value is below the threshold (< 90) or invalid.

    Thresholds:
        90–95  → Yellow
        96–98  → Orange
        99–100 → Red
    """
    if v is None:
        return None
    try:
        v = int(v)
    except Exception:
        return None
    if 90 <= v <= 95:
        return FILL_YELLOW
    if 96 <= v <= 98:
        return FILL_ORANGE
    if 99 <= v <= 100:
        return FILL_RED
    return None  # below 90 — no highlight needed


def _apply_thin(ws, r1, c1, r2, c2):
    """
    Apply a thin black border to every cell in the rectangular range
    (r1, c1) → (r2, c2).  This creates the inner grid lines of each table.
    All four sides of every cell get the same thin border.
    """
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).border = Border(
                left=THIN_SIDE, right=THIN_SIDE,
                top=THIN_SIDE,  bottom=THIN_SIDE
            )


def _apply_blue(ws, r1, c1, r2, c2):
    """
    Overlay a thick blue border on the outer edge of the rectangular range
    (r1, c1) → (r2, c2), without disturbing the inner thin borders already set.

    Strategy: read each border object, replace only the relevant side with
    THICK_BLUE, and write it back — this preserves the thin borders on the
    inner sides of corner/edge cells.
    """
    # Top and bottom edges (iterate across columns)
    for c in range(c1, c2 + 1):
        top = ws.cell(row=r1, column=c)
        bot = ws.cell(row=r2, column=c)
        top.border = Border(top=THICK_BLUE, left=top.border.left,
                            right=top.border.right, bottom=top.border.bottom)
        bot.border = Border(bottom=THICK_BLUE, left=bot.border.left,
                            right=bot.border.right, top=bot.border.top)
    # Left and right edges (iterate across rows)
    for r in range(r1, r2 + 1):
        lft = ws.cell(row=r, column=c1)
        rgt = ws.cell(row=r, column=c2)
        lft.border = Border(left=THICK_BLUE, top=lft.border.top,
                            bottom=lft.border.bottom, right=lft.border.right)
        rgt.border = Border(right=THICK_BLUE, top=rgt.border.top,
                            bottom=rgt.border.bottom, left=rgt.border.left)


def log_run(entry: dict) -> str:
    """
    Write a JSON log file for this report run to LOG_DIR.
    The filename includes a UTC timestamp so every run produces a unique file.
    Returns the path of the log file as a string.

    The log entry typically contains:
        - timestamp       : when the report was generated
        - input_summary   : counts of one_day and under_utilized entries
        - matched         : server names found in both input and master list
        - unmatched       : names in input but NOT in master (auto-add was off)
        - added           : names auto-added to master during this run
        - errors          : any exception messages
        - traceback       : full traceback if an exception occurred
    """
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    ts   = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    path = LOG_DIR / f"run_{ts}.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(entry, f, indent=2, ensure_ascii=False)
    return str(path)


# ─────────────────────────────────────────────
# MAIN REPORT BUILDING FUNCTION
# ─────────────────────────────────────────────

def build_report(data: dict, report_date_str: str, shift_str: str,
                 auto_add_new: bool = True, author: str = None):
    """
    Build the Excel server report from a parsed JSON payload.

    Parameters
    ----------
    data : dict
        Parsed JSON with two keys:
          "one_day"        : list of server dicts with avg CPU/MEM/DISK metrics
          "under_utilized" : list of server dicts for low-usage servers
    report_date_str : str
        Human-readable date, e.g. "23rd October".  Used in the title row
        and to derive the output filename.
    shift_str : str
        "Day" or "Night" — appended to the title row.
    auto_add_new : bool
        If True (default), servers in the JSON that aren't in the master list
        are automatically appended to master_servers.json for future runs.
        If False, unrecognised servers are logged but not added.

    Returns
    -------
    tuple : (xlsx_bytes, filename, log)
        xlsx_bytes : raw bytes of the .xlsx file, ready to be sent by Flask
        filename   : suggested download filename (e.g. "Server_Report_23Oct.xlsx")
        log        : the log dict that was written to disk
    """

    # Ensure master file exists before we do anything else
    _seed_master()

    # Initialise the log dict — this gets written to disk at the end regardless
    # of whether the report succeeded or failed
    log = {
        "timestamp":     datetime.now(timezone.utc).isoformat(),
        "input_summary": {},
        "matched":       [],   # servers found in both input JSON and master list
        "unmatched":     [],   # servers in input but not master (when auto_add=False)
        "added":         [],   # servers auto-added to master during this run
        "errors":        [],
    }

    # Determine the author tag for comments
    final_author = author if author else OWNER_TAG

    try:
        # ── 1. LOAD MASTER LIST ──────────────────────────────────────────────
        with open(MASTER_FILE, "r", encoding="utf-8") as f:
            master = json.load(f)

        # Build a normalised-name → original-display-name lookup for fast matching
        master_names    = [m.get("name", "") for m in master]
        master_norm_map = {norm_name(n): n for n in master_names}

        # ── 2. PARSE INPUT DATA ──────────────────────────────────────────────
        one_day = data.get("one_day", []) or []
        under   = data.get("under_utilized", []) or []

        log["input_summary"]["one_day_count"] = len(one_day)
        log["input_summary"]["under_count"]   = len(under)

        # Build one_map by MERGING all entries for the same server.
        # ChatGPT outputs one separate entry per Grafana panel — the same server can
        # appear up to three times in one_day (once with cpu_avg, once with mem_avg,
        # once with disks/disk_avg).  A plain dict comprehension would silently keep
        # only the LAST entry, losing the earlier metrics.  This loop merges all
        # fields into a single dict per server so CPU + MEM + DISK are all preserved.
        one_map: dict = {}
        for x in one_day:
            if not x.get("name"):
                continue
            key = norm_name(x.get("name", ""))
            if key not in one_map:
                one_map[key] = dict(x)   # first time we see this server: copy it in
            else:
                one_map[key].update(x)   # subsequent panels: merge new fields in

        # Build under_map by MERGING all entries for the same server, same
        # pattern as one_map above.  HOLDCO servers now appear in CPU, MEM, and
        # DISK panels separately, so a plain dict comprehension would silently
        # drop earlier entries and keep only the last one.
        under_map: dict = {}
        for x in under:
            if not x.get("name"):
                continue
            key = norm_name(x.get("name", ""))
            if key not in under_map:
                under_map[key] = dict(x)
            else:
                under_map[key].update(x)

        # ── 3. RECONCILE WITH MASTER LIST ────────────────────────────────────
        # Any server name in the incoming JSON that isn't in the master list
        # is either silently added (auto_add_new=True) or flagged (False).
        incoming_norm = set(list(one_map.keys()) + list(under_map.keys()))
        for in_name in incoming_norm:
            if in_name not in master_norm_map:
                # Recover the original (un-normalised) name from the input entry
                orig = (one_map.get(in_name) or under_map.get(in_name) or {}).get(
                    "name", in_name
                )
                if auto_add_new:
                    # Append to the in-memory master and update the lookup map
                    master.append({"name": orig})
                    master_norm_map[in_name] = orig
                    log["added"].append(orig)
                else:
                    log["unmatched"].append(orig)

        # Persist master list to disk only if new servers were added this run
        if log["added"]:
            with open(MASTER_FILE, "w", encoding="utf-8") as f:
                json.dump(master, f, indent=2, ensure_ascii=False)

        # ── 4. BUILD WORKBOOK ────────────────────────────────────────────────
        wb = Workbook()
        ws = wb.active
        ws.title = "Server Report"

        # Row 2: merged title spanning all 9 columns (A–I)
        title = f"{report_date_str} {shift_str} 2025"
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
        tc = ws.cell(row=2, column=1, value=title)
        tc.font      = Font(bold=True)
        tc.alignment = Alignment(horizontal="center")

        # Row 3: "Under-Utilized Server Resources" label over the right table (cols F–I)
        ws.merge_cells(start_row=3, start_column=6, end_row=3, end_column=9)
        rh = ws.cell(row=3, column=6, value="Under-Utilized Server Resources")
        rh.font      = Font(bold=True)
        rh.alignment = Alignment(horizontal="center")

        # Row 5: column headers for the left table (cols A–D) and right table (cols F–I)
        for i, h in enumerate(["NOS", "CPU Avg", "MEM Avg", "DISK Avg"], 1):
            ws.cell(row=5, column=i, value=h).font = Font(bold=True)
        for i, h in enumerate(["NOS", "CPU Min", "Memory", "Disk"], 6):
            ws.cell(row=5, column=i, value=h).font = Font(bold=True)

        # ── 5. WRITE LEFT TABLE (all servers, master-list order) ─────────────
        # Data starts at row 6. Every server in the master list gets a row,
        # even if it has no data this period (those cells just stay empty).
        START = 6
        r     = START
        for m in master:
            display = m.get("name", "")
            ws.cell(row=r, column=1, value=display)   # col A: server name

            nm  = norm_name(display)
            od  = one_map.get(nm)   # matching entry from the JSON, or None

            # Read CPU and MEM averages (may be absent if server not in input)
            cpu = ceil_or_none(od.get("cpu_avg")) if od else None
            mem = ceil_or_none(od.get("mem_avg")) if od else None

            # ── Disk handling ─────────────────────────────────────────────────
            # If the entry has a "disks" dict (multiple drives), take the highest
            # drive value as the reported disk figure and attach a comment that
            # lists every drive's individual reading.
            # If "disks" is absent, fall back to the simpler "disk_avg" field.
            disk_value = None
            disks = od.get("disks") if od else None

            if isinstance(disks, dict) and disks:
                # Collect valid numeric disk values
                nums = []
                for dv in disks.values():
                    try:
                        nums.append(float(dv))
                    except Exception:
                        pass  # skip non-numeric drive values gracefully
                if nums:
                    disk_value = math.ceil(max(nums))  # show worst-case drive
                    # Build comment text: one "Drive: value" line per drive + owner tag
                    comment_lines = [f"{str(k).rstrip(':')}: {v}" for k, v in disks.items()]
                    comment_lines.append(final_author)
                    ws.cell(row=r, column=4).comment = Comment(
                        "\n".join(comment_lines), final_author
                    )
            else:
                # Single-drive server — straight disk_avg value
                disk_value = ceil_or_none(od.get("disk_avg")) if od else None

            # Write CPU (col B), MEM (col C), DISK (col D) and apply colour if >= 90
            for col, val in ((2, cpu), (3, mem), (4, disk_value)):
                cell = ws.cell(row=r, column=col, value=val)
                if val is not None and val >= 90:
                    fill = pick_fill(val)
                    if fill:
                        cell.fill = fill

            # Track matched servers for the log
            if od:
                log["matched"].append(display)
            r += 1

        last_left = r - 1   # last data row in the left table

        # ── 6. WRITE RIGHT TABLE (under-utilised / internal servers) ─────────
        # These are HOLDCO (internal infrastructure) servers — AD, build, etc.
        # We use the master-list display name if the server is recognised;
        # otherwise we fall back to the name as it came from the input JSON.
        r_u = START
        for norm_k, entry in under_map.items():
            master_display = master_norm_map.get(norm_k)
            name = master_display if master_display else entry.get("name")
            ws.cell(row=r_u, column=6, value=name)                                  # col F
            ws.cell(row=r_u, column=7, value=ceil_or_none(entry.get("cpu_min")))    # col G
            ws.cell(row=r_u, column=8, value=ceil_or_none(entry.get("memory")))     # col H

            # ── Disk handling (same logic as left table) ──────────────────────
            disk_value_u = None
            disks_u = entry.get("disks")
            if isinstance(disks_u, dict) and disks_u:
                nums_u = []
                for dv in disks_u.values():
                    try:
                        nums_u.append(float(dv))
                    except Exception:
                        pass
                if nums_u:
                    disk_value_u = math.ceil(max(nums_u))
                    comment_lines_u = [f"{str(k).rstrip(':')}: {v}" for k, v in disks_u.items()]
                    comment_lines_u.append(final_author)
                    ws.cell(row=r_u, column=9).comment = Comment(
                        "\n".join(comment_lines_u), final_author
                    )
            else:
                disk_value_u = ceil_or_none(entry.get("disk"))

            ws.cell(row=r_u, column=9, value=disk_value_u)                          # col I
            r_u += 1

        # Guard: if under_map was empty, r_u is still START, so clamp to row 5
        # (the header row) so borders are applied at minimum over the header
        last_right = max(r_u - 1, 5)

        # ── 7. APPLY BORDERS ─────────────────────────────────────────────────
        # Inner thin grid first, then overlay the thick blue outer frame
        _apply_thin(ws, 5, 1, last_left,  4)   # left table
        _apply_blue(ws, 5, 1, last_left,  4)
        _apply_thin(ws, 5, 6, last_right, 9)   # right table
        _apply_blue(ws, 5, 6, last_right, 9)

        # ── 8. AUTO-SIZE COLUMNS A–I ─────────────────────────────────────────
        # Find the longest value in each column and set width accordingly.
        # Clamped between 12 (minimum) and 60 (maximum) characters + 4 padding.
        for col in range(1, 10):
            max_len = 0
            for row_idx in range(2, max(last_left, last_right) + 1):
                v = ws.cell(row=row_idx, column=col).value
                if v is None:
                    continue
                length = len(str(v))
                if length > max_len:
                    max_len = length
            ws.column_dimensions[chr(64 + col)].width = max(12, min(60, max_len + 4))

        # ── 9. DERIVE OUTPUT FILENAME ─────────────────────────────────────────
        # Try to parse the report date string (e.g. "23rd October") into a
        # clean filename like "Server_Report_23Oct.xlsx".
        # Falls back to today's date if parsing fails.
        try:
            day_str    = report_date_str.split()[0]                    # "23rd"
            digits     = "".join([c for c in day_str if c.isdigit()])  # "23"
            parts      = report_date_str.split()
            month_name = parts[1] if len(parts) > 1 else None          # "October"
            if digits and month_name:
                # Convert full month name to 3-letter abbreviation: "October" → "Oct"
                mon_abbr = __import__("datetime").datetime.strptime(
                    month_name, "%B"
                ).strftime("%b")
                filename = f"Server_Report_{int(digits)}{mon_abbr}.xlsx"
            else:
                from datetime import datetime as _dt
                filename = f"Server_Report_{_dt.now().strftime('%d%b')}.xlsx"
        except Exception:
            from datetime import datetime as _dt
            filename = f"Server_Report_{_dt.now().strftime('%d%b')}.xlsx"

        # ── 10. SERIALISE TO BYTES ────────────────────────────────────────────
        # Save the workbook into an in-memory buffer so Flask can stream it
        # directly to the browser — no temp file needed on disk.
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        xlsx_bytes = buf.read()

    except Exception as e:
        # Record the error before re-raising so the log is always written
        log["errors"].append(str(e))
        log["traceback"] = traceback.format_exc()
        log_run(log)
        raise  # bubble up to Flask which will return a 500 JSON response

    finally:
        # Always write a log — even on success — so every run is auditable.
        # Wrapped in try/except so a log failure never masks the real error.
        try:
            log_run(log)
        except Exception:
            pass

    return xlsx_bytes, filename, log
